import os
from openpyxl import load_workbook
from datetime import datetime

filepath = os.getenv('USERPROFILE')

workbook_request = load_workbook(filepath + "\\Desktop\\" + "GerarVT\\Pedido de Compra.xlsx", data_only=True)

workbook_prof = load_workbook(filepath + "\\Desktop\\" + "GerarVT\\Planilha_Base.xlsx")

sheets_request = workbook_request.sheetnames

worksheet_request = workbook_request["VT PROFESSORES"]

sheets_colab = workbook_prof.sheetnames

worksheet_prof = workbook_prof ["Vale Transporte"]

ids=[]
names=[]
totals=[]

sum_total=0

for request_cell, id_cell, quant_cell in zip (worksheet_request['C:C'], worksheet_request['A:A'], worksheet_request['E:E']):
        if request_cell.value == 2 and id_cell.value is not None and quant_cell.value != 0:
            if worksheet_request.cell(row=id_cell.row, column=1).value == worksheet_request.cell(row=id_cell.row + 1, column=1).value:
                value_cell1 = worksheet_request.cell(row=id_cell.row, column=6).value
                value_cell2 = worksheet_request.cell(row=id_cell.row + 1 , column=6).value
                total_value_cell = value_cell1 + value_cell2
                totals.append(total_value_cell)
                sum_total=sum_total+total_value_cell
                id = worksheet_request.cell(row=id_cell.row, column=1).value
                ids.append(id)
                name = worksheet_request.cell(row=id_cell.row, column=2).value
                names.append(name)
                id_cell.row = id_cell.row + 1
            elif worksheet_request.cell(row=id_cell.row, column=1).value not in ids:
                total_value_cell = worksheet_request.cell(row=id_cell.row, column=6).value
                sum_total=sum_total+total_value_cell
                totals.append(total_value_cell)
                id = worksheet_request.cell(row=id_cell.row, column=1).value
                ids.append(id)
                name = worksheet_request.cell(row=id_cell.row, column=2).value
                names.append(name)
                

for x in range (len(ids)):
    worksheet_prof.cell(row=x+5,column=1).value=ids[x]
    worksheet_prof.cell(row=x+5,column=2).value=names[x]
    worksheet_prof.cell(row=x+5,column=3).value=totals[x]

print ("Requisitantes:" ,len(ids))    
print ("Valor Total do Vale-Transporte: R$",sum_total)
current_datetime = datetime.now().strftime("%d-%m-%Y %H-%M-%S")
str_current_datetime = str(current_datetime)

workbook_prof.save(filepath + "\\Desktop\\" + "GerarVT\\Tabela Vale Transporte Professores.xlsx" + "(" + str_current_datetime + ")"+ ".xlsx")

print("Planilha de Professores Gerada")
input("Pressione Enter para continuar...")