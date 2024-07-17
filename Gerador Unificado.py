import os
from openpyxl import load_workbook
from datetime import datetime

current_datetime = datetime.now().strftime("%d-%m-%Y %H-%M-%S")
str_current_datetime = str(current_datetime)

filepath = os.getenv('USERPROFILE')

workbook_request = load_workbook(filepath + "\\Desktop\\" + "GerarVT\\Pedido de Compra.xlsx", data_only=True)

sheets = workbook_request.sheetnames

if not os.path.exists(filepath + "\\Desktop\\" + "GerarVT\\Gerados"):
    os.mkdir(filepath + "\\Desktop\\" + "GerarVT\\Gerados")
else:
    print ("A pasta já existe")


ids=[]
names=[]
totals=[]
buss=[]

metro_pass_555_quant=0
metro_pass_555_value=0
metro_pass_770_quant=0
metro_pass_770_value=0
metro_pass_980_quant=0
metro_pass_980_value=0

star_pass_510_quant=0
star_pass_510_value=0
star_pass_555_quant=0
star_pass_555_value=0
star_pass_770_quant=0
star_pass_770_value=0

fenix_pass_quant=0
fenix_pass_value=0

sum_value_pass_check = 0
sum_quant_pass_check = 0

for worksheet_request in workbook_request.worksheets[0:2]:
    os.system('cls')
    sum_value_total=0
    sum_quant_total=0
    sum_value_check=0
    sum_quant_check=0
    ids.clear()
    names.clear()
    totals.clear()
    buss.clear()
    print ("\nProcessando Informações da planilha", worksheet_request.title,"\n")    
    for request_cell, id_cell, quant_cell in zip (worksheet_request['C:C'], worksheet_request['A:A'], worksheet_request['E:E']):
        if id_cell.value is not None:
            if quant_cell.value != 0: 
                if request_cell.value in {"1 SIM", "2 SIM", 2}:
                    if worksheet_request.cell(row=id_cell.row, column=1).value == worksheet_request.cell(row=id_cell.row + 1, column=1).value and worksheet_request.cell(row=id_cell.row, column=3).value in {"1 SIM", "2 SIM", 2} and worksheet_request.cell(row=id_cell.row + 1, column=3).value in {"2 SIM", "1 SIM", 2}:
                        total_value_cell = ((worksheet_request.cell(row=id_cell.row, column=4).value * worksheet_request.cell(row=id_cell.row, column=5).value)) + ((worksheet_request.cell(row=id_cell.row + 1 , column=4).value * worksheet_request.cell(row=id_cell.row + 1, column=5).value))
                        quant_cell_total = worksheet_request.cell(row=id_cell.row, column=5).value + worksheet_request.cell(row=id_cell.row + 1 , column=5).value
                        sum_value_total=sum_value_total+total_value_cell
                        sum_quant_total=sum_quant_total+quant_cell_total
                        totals.append(total_value_cell)
                        id = worksheet_request.cell(row=id_cell.row, column=1).value
                        print ("\nID:", id)
                        ids.append(id)
                        name = worksheet_request.cell(row=id_cell.row, column=2).value
                        print ("Nome:", name)
                        names.append(name)
                        print ("Passes por Trajeto:\n", "Trajeto 1:",worksheet_request.cell(row=id_cell.row, column=5).value, "/" , "Trajeto 2:", worksheet_request.cell(row=id_cell.row + 1 , column=5).value)
                        print ("Valor Unitáro do Passe por Trajeto:\n", "Trajeto 1: R$ {:0.2f}".format(worksheet_request.cell(row=id_cell.row, column=4).value), "/" , "Trajeto 2: R$ {:0.2f}".format(worksheet_request.cell(row=id_cell.row + 1 , column=4).value))
                        print ("Valor por Trajeto:\n", "Trajeto 1: R$ {:0.2f}".format(worksheet_request.cell(row=id_cell.row, column=6).value),"/" , "Trajeto 2: R$ {:0.2f}".format(worksheet_request.cell(row=id_cell.row + 1 , column=6).value))
                        print ("Quantidade Total:" , quant_cell_total)
                        print("Valor Total: R$ {:0.2f}\n".format(total_value_cell))
                        
                    elif worksheet_request.cell(row=id_cell.row, column=1).value not in ids:
                        total_value_cell = ((worksheet_request.cell(row=id_cell.row, column=4).value * worksheet_request.cell(row=id_cell.row, column=5).value))
                        total_value_check = worksheet_request.cell(row=id_cell.row, column=6).value
                        sum_value_total=sum_value_total+total_value_cell
                        quant_cell_total = worksheet_request.cell(row=id_cell.row, column=5).value
                        sum_quant_total=sum_quant_total+quant_cell_total
                        totals.append(total_value_cell)
                        id = worksheet_request.cell(row=id_cell.row, column=1).value
                        print ("ID:", id)
                        ids.append(id)
                        name = worksheet_request.cell(row=id_cell.row, column=2).value
                        print ("Nome:", name)
                        names.append(name)
                        print ("Valor Unitáro do Passe: R$ {:0.2f}".format(worksheet_request.cell(row=id_cell.row, column=4).value))
                        print ("Quantidade Total:" , quant_cell_total)
                        print("Valor Total: R$ {:0.2f}\n".format(total_value_cell))
                
                    if worksheet_request.cell(row=id_cell.row, column=7).value in {"BIGUAÇU", "JOTUR"}:
                        if worksheet_request.cell(row=id_cell.row, column=4).value == 5.55:
                            metro_pass_555_quant = worksheet_request.cell(row=id_cell.row, column=5).value + metro_pass_555_quant
                            metro_pass_555_value = worksheet_request.cell(row=id_cell.row, column=6).value + metro_pass_555_value
                        if worksheet_request.cell(row=id_cell.row, column=4).value == 7.70:
                            metro_pass_770_quant = worksheet_request.cell(row=id_cell.row, column=5).value + metro_pass_770_quant
                            metro_pass_770_value = worksheet_request.cell(row=id_cell.row, column=6).value + metro_pass_770_value
                        if worksheet_request.cell(row=id_cell.row, column=4).value == 9.80:
                            metro_pass_980_quant = worksheet_request.cell(row=id_cell.row, column=5).value + metro_pass_980_quant
                            metro_pass_980_value = worksheet_request.cell(row=id_cell.row, column=6).value + metro_pass_980_value
                    if worksheet_request.cell(row=id_cell.row, column=7).value == "ESTRELA":    
                        if worksheet_request.cell(row=id_cell.row, column=4).value == 5.10:
                            star_pass_510_quant = worksheet_request.cell(row=id_cell.row, column=5).value + star_pass_510_quant
                            star_pass_510_value = worksheet_request.cell(row=id_cell.row, column=6).value + star_pass_510_value
                        if worksheet_request.cell(row=id_cell.row, column=4).value == 5.55:
                            star_pass_555_quant = worksheet_request.cell(row=id_cell.row, column=5).value + star_pass_555_quant
                            star_pass_555_value = worksheet_request.cell(row=id_cell.row, column=6).value + star_pass_555_value
                        if worksheet_request.cell(row=id_cell.row, column=4).value == 7.70:
                            star_pass_770_quant = worksheet_request.cell(row=id_cell.row, column=5).value + star_pass_770_quant
                            star_pass_770_value = worksheet_request.cell(row=id_cell.row, column=6).value + star_pass_770_value
                    if worksheet_request.cell(row=id_cell.row, column=7).value == "FENIX":
                            fenix_pass_quant = worksheet_request.cell(row=id_cell.row, column=5).value + fenix_pass_quant
                            fenix_pass_value = worksheet_request.cell(row=id_cell.row, column=6).value + fenix_pass_value
                    metro_pass_quant_total = metro_pass_555_quant + metro_pass_770_quant + metro_pass_980_quant
                    star_pass_quant_total = star_pass_510_quant + star_pass_555_quant + star_pass_770_quant
                    metro_pass_value_total = metro_pass_555_value + metro_pass_770_value + metro_pass_980_value
                    star_pass_value_total =  star_pass_510_value + star_pass_555_value + star_pass_770_value
        if worksheet_request.cell(row=id_cell.row, column=4).value == "TOTAL":
            sum_value_check = worksheet_request.cell(row=id_cell.row, column=6).value
            sum_quant_check = worksheet_request.cell(row=id_cell.row, column=5).value
            sum_value_pass_check =  worksheet_request.cell(row=id_cell.row, column=6).value + sum_value_pass_check
            sum_quant_pass_check =  worksheet_request.cell(row=id_cell.row, column=5).value + sum_quant_pass_check   
    print ("\nRelatório Final:")
    print ("\nRequisitantes:" ,len(ids))
    print ("Quantidade Total de Passes:",sum_quant_total)    
    print ("Valor Total dos Vales-transporte: R${:0.2f}\n".format(sum_value_total))
    if sum_value_total != sum_value_check and sum_quant_total != sum_quant_check:
        print ("A contabilização de usuário está com erro")
    else:
        print ("A contabilização de usuários está OK")
        

    user_input = input ("\nDeseja gerar a planilha de vale transporte? (s, n):")
    if user_input.lower() == "s":
        if worksheet_request.title == "VT COLABORADORES":
            workbook_base = load_workbook(filepath + "\\Desktop\\" + "GerarVT\\Planilha_Base.xlsx")
            worksheet_base = workbook_base ["Vale Transporte"]
            print ("\nTransferindo Informações para a Planilha\n")
            for x in range (len(ids)):
                worksheet_base.cell(row=x+5,column=1).value=ids[x]
                worksheet_base.cell(row=x+5,column=2).value=names[x]
                worksheet_base.cell(row=x+5,column=3).value=totals[x]
            print ("Gerando Planilha")
            workbook_base.save(filepath + "\\Desktop\\" + "GerarVT\\Gerados\\Tabela Vale Transporte Colaboradores" + "(" + str_current_datetime + ")" + ".xlsx")
            print("\nPlanilha de Colaboradores Gerada com Sucesso")
            input ("\nPressione qualquer tecla para continuar...\n")
    else:
        continue
    if worksheet_request.title == "VT PROFESSORES":
        workbook_base = load_workbook(filepath + "\\Desktop\\" + "GerarVT\\Planilha_Base.xlsx")
        worksheet_base = workbook_base ["Vale Transporte"]
        print ("\nTransferindo Informações para a Planilha\n")                
        for x in range (len(ids)):
            worksheet_base.cell(row=x+5,column=1).value=ids[x]
            worksheet_base.cell(row=x+5,column=2).value=names[x]
            worksheet_base.cell(row=x+5,column=3).value=totals[x]
        print ("Gerando Planilha")
        workbook_base.save(filepath + "\\Desktop\\" + "GerarVT\\Gerados\\Tabela Vale Transporte Professores" + "(" + str_current_datetime + ")" + ".xlsx")
        print("\nPlanilha de Professores Gerada com Sucesso\n")
        
os.system('cls')
if metro_pass_quant_total + star_pass_quant_total + fenix_pass_quant != sum_quant_pass_check and metro_pass_value_total + star_pass_value_total + fenix_pass_value != sum_value_pass_check:
    print ("A Contabilização de passes está com erro")
    print ("\nSomatório de Passes da Planilha:", sum_quant_pass_check)
    print ("\nSomatório de Passes Calculado:", metro_pass_quant_total + star_pass_quant_total + fenix_pass_quant)
else:
    print ("\nA Contabilização de passes está ok")    

print("\nGerando arquivo de passes")
with open (filepath + "\\Desktop\\" + "GerarVT\\Gerados\\Quantidade Passes.txt", "w") as file:
    data = ["**Consôrcio Fenix**", "\nQuantidade Total de Passes:", str(fenix_pass_quant), "\nValor Total: R$ {:0.2f}\n".format(fenix_pass_value), "\n**Metropolitano (Biguaçu/Jotur)**", "\nQuantidade de Passes (5,55):", str(metro_pass_555_quant), "\nValor Total (5,55): R$ {:0.2f}\n".format(metro_pass_555_value), "\nQuantidade de Passes (7,70):", str(metro_pass_770_quant), "\nValor Total (7,70): R$ {:0.2f}\n".format(metro_pass_770_value), "\nQuantidade de Passes (9,80):", str(metro_pass_980_quant), "\nValor Total (9,80): R$ {:0.2f}\n".format(metro_pass_980_value), "\nSomatório dos Passes (Metropolitano):", str(metro_pass_quant_total) , "\nSomatório do Valor dos Passes(Metropolitano): R$ {:0.2f}\n".format(metro_pass_value_total) ,"\n**Estrela**", "\nQuantidade de Passes (5,10):", str(star_pass_510_quant), "\nValor Total (5,10): R$ {:0.2f}\n".format(star_pass_510_value), "\nQuantidade de Passes (5,55):", str(star_pass_555_quant), "\nValor Total (5,55): R$ {:0.2f}\n".format(star_pass_555_value), "\nQuantidade de Passes (7,70):", str(star_pass_770_quant), "\nValor Total (7,70): R$ {:0.2f}\n".format(star_pass_770_value), "\nSomatório dos Passes (Estrela):", str(star_pass_quant_total) , "\nSomatório do Valor dos Passes(Estrela): R$ {:0.2f}\n".format(star_pass_value_total)]
    for line in data:
        file.write(line)
print("\nArquivo de passes gerado")
input("\nPressione qualquer tecla para encerrar a aplicação...")